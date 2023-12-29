import os
import time
import pandas as pd
import matplotlib.pyplot as plt

from openpyxl import Workbook

logs_folder = 'logs'
log_data = []

# 遍历 logs 文件夹下的所有文件
for filename in os.listdir(logs_folder):
    if filename.endswith('.log'):
        # 构建日志文件的完整路径
        log_filepath = os.path.join(logs_folder, filename)

        with open(log_filepath, 'r') as log_file:
            recognize_lines = []
            for line in log_file:
                if 'resultDishCount' in line:
                    # 提取 resultDishCount 后的值
                    result_dish_count = line.split('resultDishCount: ')[1].split(' ')[0]
                    if int(result_dish_count) <= 8:
                        data = {'resultDishCount': result_dish_count}
                        recognize_lines.clear()
                        next(log_file, None)
                        next_line = next(log_file, None)
                        if next_line and ' recognize: ' in next_line:
                            recognize_lines.append(next_line)
                            # print(recognize_lines)
                            recognize_time1 = recognize_lines[0].split('recognize: ')[1].split('ms')[0]
                            data['recognizeTime'] = recognize_time1
                            log_data.append(data)
                        else:
                            next_line = next(log_file, None)
                            if next_line and ' recognize: ' in next_line:
                                recognize_lines.append(next_line)
                                # print(recognize_lines)
                                recognize_time1 = recognize_lines[0].split('recognize: ')[1].split('ms')[0]
                                data['recognizeTime'] = recognize_time1
                                log_data.append(data)

# 创建工作簿和工作表
workbook = Workbook()
worksheet = workbook.active

# 添加表头
worksheet['A1'] = 'resultDishCount'
worksheet['B1'] = 'recognizeTime'

# 写入数据到单元格
for i, data in enumerate(log_data):
    row = i + 2  # 从第二行开始写入数据，因为第一行是表头
    worksheet.cell(row, column=1, value=data.get('resultDishCount'))
    worksheet.cell(row, column=2, value=data.get('recognizeTime'))

# 保存 Excel 文件
output_filename = 'excel/' + time.strftime("%Y-%m-%d_%H-%M-%S") + 'out.xlsx'
workbook.save(output_filename)
print("created " + output_filename)

# 读取 Excel 文件并绘制透视图和条形图
df = pd.read_excel(output_filename)
pivot_table = pd.pivot_table(df, values='recognizeTime', index='resultDishCount', aggfunc='mean')
a = pivot_table.plot(y='recognizeTime', kind='bar')
print(a)
plt.savefig('excel/bar_chart.png')
# 将最后生成的数据写到output_filename的sheet2中
worksheet2 = workbook.create_sheet(title='平均值')
worksheet2['A1'] = 'resultDishCount（个）'
worksheet2['B1'] = 'recognizeTime（ms）'
for i, (index, row) in enumerate(pivot_table.iterrows()):
    worksheet2.cell(row=i + 2, column=1, value=index)
    worksheet2.cell(row=i + 2, column=2, value=row['recognizeTime'])
from openpyxl.drawing.image import Image

img = Image('excel/bar_chart.png')
worksheet2.add_image(img, 'D1')
workbook.save(output_filename)
